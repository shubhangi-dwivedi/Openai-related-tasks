import openai
import json
import openpyxl
import string

openai.api_key_path = "apikey.txt"

def Convert2(string):
    li = list(string.split('\n'))
    return li

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = 'Quote'
ws.cell(row=1, column=2).value = 'Meta data/Keywords'

inputExcelFile = "goodreads_quotes.xlsx"

newWorkbook = openpyxl.load_workbook(inputExcelFile)

sheet = newWorkbook["Sheet1"]
row_count = sheet.max_row

for row_data in range(1, row_count + 1):

    rk = sheet.cell(row=row_data, column=2).value

    response = openai.Completion.create(
        model="text-curie-001",
        prompt='"' + rk + '" '+'find 25-30 comma separated keywords related to above quote',
        temperature=0.2,
        max_tokens=600,
        top_p=0,
        frequency_penalty=0,
        presence_penalty=0
    )

    s1 = json.dumps(response)
    json_object = json.loads(s1)
    output = ""
    li=[]

    for each in json_object['choices']:
        x = each['text']
        output = x
        output = output.strip("\n")
        output = output.replace('"', "")
        li2 = Convert2(output)

        for x2 in li2:
            x2= x2.replace("-"," ")
            x2=x2.replace("\n"," ")
            x2=x2.replace('"','')

            if len(x2)==0:
                continue

            if len(x2) ==1 and x2[0]=='.':
                x2=x2.replace('.','')
                x2 = x2.replace('"','')
                continue

            if len(x2) >= 4 and (x2[1]== '.' or x2[2]=='.'):
                ind = x2.index(".")
                x2 = x2[ind + 1:]
                x2 = x2.replace('"','')
                li.append(x2)
                continue

            if len(x2)>=0:
                    #x2=x2.split(',')
                li.append(x2)

                    #print(output)
            else :
                li.append(x2)

        print(li)
        temp2=str(li)
        temp2=temp2.replace("[","")
        temp2 = temp2.replace("]", "")
        temp2 = temp2.replace('"','')
        temp2 = temp2.replace("'","")

        print(temp2)
        l = []
        l.append(rk)
        l.append(temp2)

        data = tuple(l)
                        # print(l)
        ws.append(data)
        wb.save('quotes_meta-data.xlsx')

