import openai
import json
import openpyxl
import string

openai.api_key_path = "apikey.txt"


def Convert1(string):
    li = list(string.split(","))
    return li


def Convert2(string):
    li = list(string.split('\n'))
    return li


wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = 'Industry'
ws.cell(row=1, column=2).value = 'Subcategory'
ws.cell(row=1, column=3).value = 'Related Keyword'
ws.cell(row=1, column=4).value = 'Ad. slogan'
ws.cell(row=1, column=5).value = 'Target audience'

inputExcelFile = "data.xlsx"

newWorkbook = openpyxl.load_workbook(inputExcelFile)

sheet = newWorkbook["Sheet1"]
row_count = sheet.max_row

temp = 0
for row_data in range(1, row_count + 1):

    if temp != 0:
        print(row_data)
        industry = sheet.cell(row=row_data, column=1).value
        subcategory = sheet.cell(row=row_data, column=2).value
        rk = sheet.cell(row=row_data, column=3).value
        target_audi = sheet.cell(row=row_data, column=4).value

        word = rk
        li = Convert1(word)

        rk1 = ""

        for x in li:
            rk1 = x
            response = openai.Completion.create(
                model="text-curie-001",
                prompt="2 advertisement slogans on " + x,
                temperature=0,
                max_tokens=4000,
                top_p=1,
                frequency_penalty=0,
                presence_penalty=0
            )

            s1 = json.dumps(response)
            json_object = json.loads(s1)
            output = ""

            for each in json_object['choices']:
                x = each['text']
                output = x

                output = output.strip("\n")
                output = output.replace('"', "")
                li2 = Convert2(output)
                print(li2)

                for x2 in li2:
                    ind = x2.index(".")
                    x2 = x2[ind + 1:]
                    print(x2)
                    l = []
                    l.append(industry)
                    l.append(subcategory)
                    l.append(rk1)
                    l.append(x2)
                    l.append(target_audi)

                    data = tuple(l)
                    # print(l)
                    ws.append(data)
        wb.save('ad_slogans.xlsx')

    temp += 1
    # nonalpha = string.digits + string.punctuation + string.whitespace
    # print(output)
